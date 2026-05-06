"""
Creates data/data_template.xlsx — the input file used each year.
Run once: python build_data_template.py
Then fill in the yellow cells with real data and run generate_report.py.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

DG  = "1A4731"   # dark green
GO  = "F5B800"   # gold
LG  = "EAF4EE"   # light green (formula)
YEL = "FFFF99"   # yellow (user input)
GRY = "D9D9D9"   # grey (locked / not used)
WHT = "FFFFFF"
BLK = "000000"
BLU = "0070C0"

thin = Side(style="thin", color="AAAAAA")
B    = Border(left=thin, right=thin, top=thin, bottom=thin)
med  = Side(style="medium", color=DG)
Bmed = Border(left=thin, right=thin, top=thin, bottom=med)

def bg(c): return PatternFill("solid", fgColor=c)

def H(ws, r, c, v, w=WHT, fill=DG, align="center", sz=11):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Calibri", bold=True, color=w, size=sz)
    cell.fill = bg(fill)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = B
    return cell

def I(ws, r, c, v=None, note=None):
    """User input cell — yellow background, blue text"""
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Calibri", color=BLU, size=11)
    cell.fill = bg(YEL)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = B
    if note:
        cell.comment = None  # openpyxl comments need extra lib; skip
    return cell

def F(ws, r, c, formula, fill=LG, fmt=None, align="center"):
    """Formula cell — light green"""
    cell = ws.cell(row=r, column=c, value=formula)
    cell.font = Font(name="Calibri", color=BLK, size=11)
    cell.fill = bg(fill)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = B
    if fmt: cell.number_format = fmt
    return cell

def L(ws, r, c, v, bold=False, align="left", fill=WHT, color="333333"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Calibri", bold=bold, color=color, size=11)
    cell.fill = bg(fill)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = B
    return cell

def banner(ws, text, end_col="F"):
    ws.row_dimensions[1].height = 36
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name="Calibri", bold=True, color=WHT, size=12)
    c.fill = bg(DG)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(f"A1:{end_col}1")

def note(ws, r, text, end_col="F"):
    ws.row_dimensions[r].height = 32
    c = ws.cell(row=r, column=1, value=text)
    c.font = Font(name="Calibri", italic=True, color="555555", size=9)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(f"A{r}:{end_col}{r}")

# ── SHEET 0: INSTRUCTIONS ────────────────────────────────────────
ws0 = wb.active
ws0.title = "INSTRUCTIONS"
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 22
ws0.column_dimensions["B"].width = 68

banner(ws0, "Hankamer School of Business  |  Annual Report — Data Entry Workbook", "B")

steps = [
    ("HOW TO GENERATE THE REPORT", None, True),
    ("Step 1", "Fill in all YELLOW cells in each sheet below with the new year's data.", False),
    ("Step 2", "Save this file as data/hankamer_data_AY2627.xlsx (rename for the year).", False),
    ("Step 3", "Open a terminal in the project folder and run:  python generate_report.py --input data/hankamer_data_AY2627.xlsx", False),
    ("Step 4", "The finished PowerPoint will appear in the output/ folder, ready to present.", False),
    ("", None, False),
    ("SHEET GUIDE", None, True),
    ("config",           "Academic year label, session count, visitor estimate, states count — drives slide 1 headlines.", False),
    ("attendance",       "Monthly headcount per semester — drives the bar chart on slide 2.", False),
    ("grad_year",        "Student count by graduation class — drives the pie chart on slide 2.", False),
    ("top10_states",     "Top 10 states by count (sorted largest→smallest) — drives the bar chart on slide 4.", False),
    ("regional",         "Regional breakdown totals — drives the table and callout boxes on slide 3.", False),
    ("companion",        "Who joins students (qualitative frequency labels) — drives the table on slide 4.", False),
    ("bullets",          "Editable text for overview, recommendations, and engagement insights — drives all bullet text.", False),
    ("", None, False),
    ("COLOUR GUIDE", None, True),
    ("Yellow cells",  "→  Enter your data here. These are the ONLY cells you should change.", False),
    ("Green cells",   "→  Auto-calculated formulas. Do not edit.", False),
    ("White cells",   "→  Labels / structure. Do not edit.", False),
]

r = 2
for item in steps:
    is_hdr, left, right = item[0] if isinstance(item[0], bool) else False, item[0], item[1]
    is_hdr = item[2]
    ws0.row_dimensions[r].height = 26 if not is_hdr else 30
    if is_hdr:
        if left:
            c = ws0.cell(row=r, column=1, value=left)
            c.font = Font(name="Calibri", bold=True, color=WHT, size=11)
            c.fill = bg(DG)
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws0.merge_cells(f"A{r}:B{r}")
        else:
            ws0.row_dimensions[r].height = 10
    else:
        cl = ws0.cell(row=r, column=1, value=left)
        cl.font = Font(name="Calibri", bold=True, color=DG, size=11)
        cl.alignment = Alignment(horizontal="left", vertical="center")
        if right:
            cr = ws0.cell(row=r, column=2, value=right)
            cr.font = Font(name="Calibri", color="333333", size=11)
            cr.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    r += 1

# ── SHEET 1: config ──────────────────────────────────────────────
ws1 = wb.create_sheet("config")
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 28
ws1.column_dimensions["C"].width = 40

banner(ws1, "config  |  High-level numbers that appear on slide 1 and throughout the deck.", "C")

H(ws1, 2, 1, "Field", align="left")
H(ws1, 2, 2, "Value (edit yellow)")
H(ws1, 2, 3, "Where used in PowerPoint", align="left")

config_rows = [
    ("academic_year",        "2025–2026",  "Slide 1 title, slide 2/3 headers, footnotes"),
    ("fall_label",           "Fall 2025",  "Bar chart legend, slide 1 stat box"),
    ("spring_label",         "Spring 2026","Bar chart legend, slide 1 stat box"),
    ("fall_students",        249,          "Slide 1 stat box (auto if attendance sheet filled)"),
    ("spring_students",      389,          "Slide 1 stat box (auto if attendance sheet filled)"),
    ("total_sessions",       117,          "Slide 1 stat box"),
    ("est_total_visitors",   "~1,600",     "Slide 1 stat box, family engagement bullet"),
    ("states_represented",   38,           "Slide 1 stat box, slide 3 header"),
    ("top2_classes_pct",     "92%",        "Slide 2 callout box (Class of 2026 & 2027 = X%)"),
    ("top2_classes_label",   "Class of 2026 & 2027", "Slide 2 callout box label"),
    ("visitor_multiplier",   "2.5×",       "Slide 4 family engagement bullet"),
    ("footnote_slide4",      "Visitor estimate based on ~2.5× the student headcount; companion data is self-reported at time of registration. Companion frequency reflects qualitative patterns, not exact counts.", "Slide 4 footnote"),
        ("footer_text",          "Hankamer School of Business  ·  Baylor University  ·  Information Session Registration Log, AY 2025–2026", "Slide 4 footer"),
]

for i, (field, value, where) in enumerate(config_rows):
    row = i + 3
    ws1.row_dimensions[row].height = 24
    fill = WHT if i % 2 == 0 else "F5F5F5"
    L(ws1, row, 1, field, bold=True, fill=fill)
    I(ws1, row, 2, value)
    L(ws1, row, 3, where, fill=fill)

note(ws1, len(config_rows) + 4,
     "⚠  fall_students and spring_students will be auto-calculated from the attendance sheet if you fill that in. "
     "Only override manually if attendance sheet is incomplete.", "C")

# ── SHEET 2: attendance ──────────────────────────────────────────
ws2 = wb.create_sheet("attendance")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 16
ws2.column_dimensions["E"].width = 30

banner(ws2, "attendance  |  Monthly student headcount. Leave cells blank (not 0) for months outside that semester.", "E")

H(ws2, 2, 1, "Month")
H(ws2, 2, 2, "Fall count")
H(ws2, 2, 3, "Spring count")
H(ws2, 2, 4, "Row total")
H(ws2, 2, 5, "Notes", align="left")

months = [
    ("Aug",  "Fall",   26,   None, ""),
    ("Sep",  "Fall",   53,   None, ""),
    ("Oct",  "Fall",   73,   None, ""),
    ("Nov",  "Fall",   62,   None, ""),
    ("Dec",  "Fall",   35,   None, "Fall ends"),
    ("Jan",  "Spring", None, 19,   "Spring begins"),
    ("Feb",  "Spring", None, 89,   ""),
    ("Mar",  "Spring", None, 154,  "Peak month"),
    ("Apr",  "Spring", None, 127,  ""),
]

for i, (mon, sem, fv, sv, n) in enumerate(months):
    r = i + 3
    ws2.row_dimensions[r].height = 22
    alt = WHT if i % 2 == 0 else "F5F5F5"
    # Month label — user must also enter the full label e.g. "Aug '25"
    I(ws2, r, 1, f"{mon} '{str(25 + (0 if sem=='Fall' else 1))[2:]}")
    if fv is not None:
        I(ws2, r, 2, fv)
    else:
        c = ws2.cell(row=r, column=2)
        c.fill = bg(GRY); c.border = B
    if sv is not None:
        I(ws2, r, 3, sv)
    else:
        c = ws2.cell(row=r, column=3)
        c.fill = bg(GRY); c.border = B
    F(ws2, r, 4, f'=IF(B{r}<>"",B{r},IF(C{r}<>"",C{r},""))')
    L(ws2, r, 5, n)

ws2.row_dimensions[12].height = 26
H(ws2, 12, 1, "Semester total")
F(ws2, 12, 2, "=SUM(B3:B11)")
F(ws2, 12, 3, "=SUM(C3:C11)")
F(ws2, 12, 4, "=B12+C12")
L(ws2, 12, 5, "← must equal grad_year total", fill=LG)

ws2.row_dimensions[13].height = 26
H(ws2, 13, 1, "Grand total (AY)")
F(ws2, 13, 2, "=B12+C12")
ws2.merge_cells("B13:D13")

ws2.row_dimensions[15].height = 22
chk = ws2.cell(row=15, column=1, value="Cross-check vs grad_year:")
chk.font = Font(name="Calibri", bold=True, color=DG, size=11)
chk.alignment = Alignment(horizontal="left", vertical="center")
ws2.merge_cells("A15:B15")
F(ws2, 15, 3, '=IF(B13=grad_year!B11,"✔ MATCH","✗ MISMATCH — check totals")',
  align="left")
ws2.merge_cells("C15:E15")

# ── SHEET 3: grad_year ───────────────────────────────────────────
ws3 = wb.create_sheet("grad_year")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 16
ws3.column_dimensions["C"].width = 14
ws3.column_dimensions["D"].width = 30

banner(ws3, "grad_year  |  Student count by graduation class (combined AY). Drives the pie chart on slide 2.", "D")

H(ws3, 2, 1, "Graduation class", align="left")
H(ws3, 2, 2, "Student count")
H(ws3, 2, 3, "% of total")
H(ws3, 2, 4, "Notes", align="left")

grad = [
    ("Class of 2026", 360, "Seniors"),
    ("Class of 2027", 228, "Juniors"),
    ("Class of 2028", 26,  "Sophomores"),
    ("Transfer",      15,  "Any year"),
    ("Class of 2029+", 9,  "Freshmen / future"),
]
for i, (cls, cnt, n) in enumerate(grad):
    r = i + 3
    ws3.row_dimensions[r].height = 22
    L(ws3, r, 1, cls, bold=True)
    I(ws3, r, 2, cnt)
    F(ws3, r, 3, f"=IF(B11>0,B{r}/B11,\"\")", fmt="0%")
    L(ws3, r, 4, n)

ws3.row_dimensions[11].height = 26
H(ws3, 11, 1, "Total", align="left")
F(ws3, 11, 2, "=SUM(B3:B7)")
F(ws3, 11, 3, "=SUM(C3:C7)", fmt="0%")

ws3.row_dimensions[13].height = 22
chk = ws3.cell(row=13, column=1, value="Cross-check vs attendance:")
chk.font = Font(name="Calibri", bold=True, color=DG, size=11)
chk.alignment = Alignment(horizontal="left", vertical="center")
ws3.merge_cells("A13:B13")
F(ws3, 13, 3, '=IF(B11=attendance!B13,"✔ MATCH","✗ MISMATCH")', align="left")
ws3.merge_cells("C13:D13")

# ── SHEET 4: top10_states ────────────────────────────────────────
ws4 = wb.create_sheet("top10_states")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 10
ws4.column_dimensions["B"].width = 16
ws4.column_dimensions["C"].width = 20
ws4.column_dimensions["D"].width = 28

banner(ws4, "top10_states  |  Enter states LARGEST first (TX at top). Script sorts ascending for the chart automatically.", "D")

H(ws4, 2, 1, "State abbr.")
H(ws4, 2, 2, "Student count")
H(ws4, 2, 3, "Full state name", align="left")
H(ws4, 2, 4, "Notes", align="left")

states = [
    ("TX", 414, "Texas",      "Home state"),
    ("CA", 48,  "California", "#2 non-TX"),
    ("CO", 19,  "Colorado",   ""),
    ("IL", 15,  "Illinois",   ""),
    ("AZ", 14,  "Arizona",    ""),
    ("WA", 10,  "Washington", ""),
    ("KS", 9,   "Kansas",     ""),
    ("MN", 8,   "Minnesota",  ""),
    ("NY", 8,   "New York",   ""),
    ("MO", 7,   "Missouri",   ""),
]
for i, (st, cnt, name, n) in enumerate(states):
    r = i + 3
    ws4.row_dimensions[r].height = 22
    I(ws4, r, 1, st)
    I(ws4, r, 2, cnt)
    I(ws4, r, 3, name)
    L(ws4, r, 4, n)

note(ws4, 14,
     "Enter states in LARGEST→SMALLEST order (TX first). The script will reverse them for the chart (which reads bottom→top).", "D")

# ── SHEET 5: regional ────────────────────────────────────────────
ws5 = wb.create_sheet("regional")
ws5.sheet_view.showGridLines = False
ws5.column_dimensions["A"].width = 22
ws5.column_dimensions["B"].width = 14
ws5.column_dimensions["C"].width = 14
ws5.column_dimensions["D"].width = 30

banner(ws5, "regional  |  Student counts by region. Drives the table and callout boxes on slide 3.", "D")

H(ws5, 2, 1, "Region", align="left")
H(ws5, 2, 2, "Students")
H(ws5, 2, 3, "Share %")
H(ws5, 2, 4, "States included", align="left")

regions = [
    ("Texas",             414, "TX"),
    ("West / SW",         103, "CA, CO, AZ, WA, NM, NV, UT, OR, ID, WY, MT, HI, AK"),
    ("Midwest",           57,  "IL, MN, MO, KS, IN, OH, MI, WI, IA, SD, ND, NE"),
    ("South / Southeast", 41,  "TN, NC, GA, FL, AL, AR, LA, SC, VA, MD, DE, OK"),
    ("Northeast",         21,  "NY, MA, CT, NJ, PA, NH, VT, ME, RI"),
    ("International",     1,   "Torreón, Mexico"),
]
for i, (reg, cnt, states_list) in enumerate(regions):
    r = i + 3
    ws5.row_dimensions[r].height = 22
    L(ws5, r, 1, reg, bold=True)
    I(ws5, r, 2, cnt)
    F(ws5, r, 3, f"=IF(B10>0,B{r}/B10,\"\")", fmt="0%")
    L(ws5, r, 4, states_list)

ws5.row_dimensions[10].height = 26
H(ws5, 10, 1, "Resolved total", align="left")
F(ws5, 10, 2, "=SUM(B3:B8)")
F(ws5, 10, 3, "=SUM(C3:C8)", fmt="0%")

# Callout auto-calc
ws5.row_dimensions[12].height = 28
H(ws5, 12, 1, "Callout values (auto)", fill=GO, w=DG, align="left")
ws5.merge_cells("A12:D12")

H(ws5, 13, 1, "Callout", align="left")
H(ws5, 13, 2, "Value")
H(ws5, 13, 3, "Display text")
H(ws5, 13, 4, "Used in PowerPoint", align="left")

callouts = [
    ("Texas %",        "=ROUND(B3/B10,2)", '=TEXT(B14,"0%")&" from Texas"',    "Slide 3 left callout box"),
    ("Out-of-state %", "=ROUND((B10-B3-B8)/B10,2)", '=TEXT(B15,"0%")&" out-of-state"', "Slide 3 centre callout box"),
    ("International",  "=B8",             '=B16&" international"',             "Slide 3 right callout box"),
]
for i, (label, val_f, disp_f, n) in enumerate(callouts):
    r = i + 14
    ws5.row_dimensions[r].height = 22
    L(ws5, r, 1, label, bold=True)
    F(ws5, r, 2, val_f, fmt="0%")
    F(ws5, r, 3, disp_f, align="left")
    L(ws5, r, 4, n)

# ── SHEET 6: companion ───────────────────────────────────────────
ws6 = wb.create_sheet("companion")
ws6.sheet_view.showGridLines = False
ws6.column_dimensions["A"].width = 24
ws6.column_dimensions["B"].width = 18
ws6.column_dimensions["C"].width = 30

banner(ws6, "companion  |  Who joins students? Drives the qualitative frequency table on slide 4.", "C")

H(ws6, 2, 1, "Companion type", align="left")
H(ws6, 2, 2, "Frequency label")
H(ws6, 2, 3, "Allowed values", align="left")

companion = [
    ("Mother (solo)",      "Very common"),
    ("Father (solo)",      "Common"),
    ("Both parents",       "Very common"),
    ("Sibling(s)",         "Occasional"),
    ("Extended family",    "Rare"),
    ("Friends / classmates","Rare"),
    ("Student alone",      "Uncommon"),
]
allowed = "Very common / Common / Occasional / Rare / Uncommon"
for i, (comp, freq) in enumerate(companion):
    r = i + 3
    ws6.row_dimensions[r].height = 22
    L(ws6, r, 1, comp, bold=True)
    I(ws6, r, 2, freq)
    L(ws6, r, 3, allowed if i == 0 else "")

# ── SHEET 7: bullets ─────────────────────────────────────────────
ws7 = wb.create_sheet("bullets")
ws7.sheet_view.showGridLines = False
ws7.column_dimensions["A"].width = 22
ws7.column_dimensions["B"].width = 74

banner(ws7, "bullets  |  All editable text blocks. Each cell is one bullet or sentence in the PowerPoint.", "B")

sections = [
    ("SLIDE 1 — OVERVIEW (6 sentences)", None, True),
    ("overview_1", "638 students attended Hankamer info sessions across AY 2025–2026 — 249 in fall, 389 in spring."),
    ("overview_2", "117 sessions hosted across the year averaged just ~5 students each — a high-touch, personalized experience rarely found at scale."),
    ("overview_3", "Spring 2026 outpaced Fall 2025 by 56%, driven by senior application timelines and growing program interest."),
    ("overview_4", "Most students brought 1–2 family members, generating an estimated ~1,600 total campus visitors across the year."),
    ("overview_5", "38 states represented across both semesters — visitors came from coast to coast, including VT, OR, NY, and NH."),
    ("overview_6", "Class of 2026 and 2027 together made up 92% of all visitors, with a healthy mix of near-term and imminent prospects."),
    ("SLIDE 1 — RECOMMENDATIONS (3 bullets)", None, True),
    ("rec_1", "Scale spring capacity — March and April are peak months; additional session dates or a larger venue may be needed."),
    ("rec_2", "Target high-potential out-of-state clusters: Kansas, Colorado, Illinois, Arkansas, and Oklahoma show strong interest."),
    ("rec_3", "Design for the family — ~40% of visits included both parents; messaging should address ROI, outcomes, and campus culture."),
    ("SLIDE 2 — BAR CHART FOOTNOTE", None, True),
    ("slide2_footnote", "Both semesters shown. Fall 2025: 249 students (Aug–Dec). Spring 2026: 389 students (Jan–Apr). Total: 638."),
    ("SLIDE 3 — GEOGRAPHIC BULLETS (2 bullets)", None, True),
    ("geo_bullet_1", "CA (#2, 48), CO (19), IL (15), AZ (14), WA (10), KS (9) are top non-TX states."),
    ("geo_bullet_2", "One student listed Torreón, Mexico — nascent international reach."),
    ("SLIDE 3 — GEO FOOTNOTE", None, True),
    ("slide3_footnote", "AY 2025–2026 combined. Some records had missing/ambiguous state data. Regional counts sum to 637 of 638 total students; 1 student had an unresolvable state entry."),
    ("SLIDE 4 — FAMILY ENGAGEMENT BULLETS (3 bullets)", None, True),
    ("engage_1", "~1,600 total campus visitors estimated across AY 2025–2026 — nearly 2.5× the student headcount."),
    ("engage_2", "~40% of visits included both parents, making this as much a family event as a student one."),
    ("engage_3", "Younger siblings (Class of 2028–2029+) are forming early impressions — a long-lead pipeline."),
]

r = 2
for item in sections:
    if item[2] if len(item) == 3 else False:
        ws7.row_dimensions[r].height = 28
        c = ws7.cell(row=r, column=1, value=item[0])
        c.font = Font(name="Calibri", bold=True, color=WHT, size=11)
        c.fill = bg(DG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws7.merge_cells(f"A{r}:B{r}")
    else:
        ws7.row_dimensions[r].height = 38
        key, val = item[0], item[1]
        L(ws7, r, 1, key, bold=True, color=DG)
        cell = ws7.cell(row=r, column=2, value=val)
        cell.font = Font(name="Calibri", color=BLU, size=11)
        cell.fill = bg(YEL)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = B
    r += 1

wb.save("data/data_template.xlsx")
print("✔ data/data_template.xlsx created")
